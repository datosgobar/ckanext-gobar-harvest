from __future__ import absolute_import

import io
import json
import logging
import hashlib
import requests
import openpyxl

from ckan.plugins import toolkit
from ckanext.harvest.model import HarvestObject, HarvestGatherError

from .base import HarvesterBase

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Mapa de columnas del XLSX → field_name del schema CKAN/datos.gob.ar
#
# Los encabezados del Excel corresponden a la display_property del perfil
# DCAT-AP de datos.gob.ar (ej. "dct:title"). Cuando un campo no tiene
# display_property se usa su field_name directamente (ej. "name").
# ---------------------------------------------------------------------------

# Pestaña "Datasets"
DATASET_COLUMN_MAP = {
    "dataset_identifier":     "id",       # clave primaria (UUID)
    "dct:title":              "dataset_title",
    "name":                   "name",
    "dct:publisher":          "owner_org",
    "dataset_publisher_mbox": "dataset_publisher_mbox",
    "dct:description":        "dataset_description",
    "dct:issued":             "dataset_issued",
    "dct:modified":           "dataset_modified",
    "adms:status":            "dataset_status",
    "dct:theme":              "dataset_superTheme",
    "dcat:theme":             "dataset_theme",
    "dataset_keywords":       "dataset_keywords",
    "dcatap:hdvCategory":     "dataset_hvdcategory",
    "dct:accrualPeriodicity": "dataset_accrualPeriodicity",
    "temporal_start":           "temporal_start",
    "temporal_end":           "temporal_end",
    "dct:spatial":                "spatial",
    "spatial_uri":            "spatial_uri",
    "spatial_coverage":       "spatial_coverage",
    "license_id":             "license_id",
    "dataset_source":          "dataset_source",
}

# Pestaña "Distributions"
RESOURCE_COLUMN_MAP = {
    "dataset_identifier":            "id",       # FK → Datasets
    "distribution_identifier":       "distribution_identifier",  # clave propia
    "url":                           "url",
    "name":                          "name",
    "description":                   "description",
    "format":                        "format",
    "dcat:mediaType":                "distribution_mediaType",
    "dct:type":                      "distribution_type",
    "cnt:characterEncoding":         "distribution_character_set",
    "dcat:spatialResolutionInMeters":"distribution_scale",
    "dct:conformsTo":                "distribution_projection",
    "distribution_iso19115_url":     "distribution_iso19115_url",
    "distribution_wfs_url":            "distribution_wfs_url",
}

# Nombre de las pestañas (configurable via JSON de configuración)
DEFAULT_DATASET_SHEET     = "Datasets"
DEFAULT_DISTRIBUTION_SHEET = "Distributions"

# Columnas obligatorias: filas sin estos valores se descartan
DATASET_REQUIRED     = {"dataset_title", "owner_org", "dataset_description"}
DISTRIBUTION_REQUIRED = {"url"}


class XLSXHarvester(HarvesterBase):
    """
    Harvester para archivos Excel (.xlsx) con el esquema de metadatos
    datos.gob.ar (perfil DCAT-AP).

    Estructura esperada del archivo:
      - Pestaña "Datasets"      : una fila por dataset; encabezados = display_property
      - Pestaña "Distributions" : una fila por recurso; incluye dataset_identifier (FK)

    Configuración opcional (JSON en el campo "Configuration" del formulario):
    {
        "dataset_sheet":       "Datasets",       // nombre de la hoja de datasets
        "distribution_sheet":  "Distributions",  // nombre de la hoja de distribuciones
        "skip_rows":           0,                // filas extra a saltar tras el encabezado
        "default_owner_org":   "mi-org",         // org. por defecto si la columna está vacía
        "user_agent":          "mi-portal/1.0"
    }
    """

    config = None

    # ------------------------------------------------------------------ info --

    def info(self):
        return {
            "name": "xlsx",
            "title": "XLSX (datos.gob.ar)",
            "description": (
                "Harvests remote XLSX files following the datos.gob.ar DCAT-AP schema. "
                "Expects a 'Datasets' sheet and a 'Distributions' sheet."
            ),
            "form_config_interface": "Text",
        }

    # --------------------------------------------------------- validate_config --

    def validate_config(self, config_str):
        if not config_str:
            return config_str
        try:
            config = json.loads(config_str)
        except ValueError as e:
            raise ValueError("La configuración debe ser un JSON válido: %s" % e)
        for key in ("dataset_sheet", "distribution_sheet", "default_owner_org", "user_agent"):
            if key in config and not isinstance(config[key], str):
                raise ValueError("%s debe ser un string" % key)
        if "skip_rows" in config and not isinstance(config["skip_rows"], int):
            raise ValueError("skip_rows debe ser un entero")
        return config_str

    # ------------------------------------------------------ get_original_url --

    def get_original_url(self, harvest_object_id):
        obj = HarvestObject.get(harvest_object_id)
        return obj.source.url if obj else None

    # ---------------------------------------------------------- gather_stage --

    def gather_stage(self, harvest_job):
        log.info("XLSXHarvester gather_stage: %s", harvest_job.source.url)

        self._set_config(harvest_job.source.config)
        source_url = harvest_job.source.url.strip()

        # Descargar el workbook
        try:
            wb = self._fetch_workbook(source_url)
        except Exception as e:
            self._save_gather_error("No se pudo descargar/leer el XLSX: %s" % e, harvest_job)
            return []

        # Leer ambas hojas
        skip = self.config.get("skip_rows", 0)

        ds_sheet_name   = self.config.get("dataset_sheet", DEFAULT_DATASET_SHEET)
        dist_sheet_name = self.config.get("distribution_sheet", DEFAULT_DISTRIBUTION_SHEET)

        try:
            ds_rows   = self._read_sheet(wb, ds_sheet_name,   DATASET_COLUMN_MAP,   skip)
            dist_rows = self._read_sheet(wb, dist_sheet_name, RESOURCE_COLUMN_MAP, skip)
        except Exception as e:
            self._save_gather_error("Error leyendo hojas del XLSX: %s" % e, harvest_job)
            return []

        if not ds_rows:
            self._save_gather_error("La hoja de datasets está vacía", harvest_job)
            return []

        # Agrupar distribuciones por dataset_identifier
        dist_by_dataset = {}
        for d in dist_rows:
            ds_id = d.get("dataset_identifier", "").strip()
            if ds_id:
                dist_by_dataset.setdefault(ds_id, []).append(d)

        # Crear un HarvestObject por dataset
        object_ids = []
        for i, row in enumerate(ds_rows):
            ds_id = row.get("dataset_identifier", "").strip()

            # Validar campos obligatorios
            missing = [f for f in DATASET_REQUIRED if not row.get(f, "").strip()]
            if missing:
                log.warning("Fila de dataset %d omitida: faltan %s", i + 2, missing)
                continue

            # Adjuntar las distribuciones correspondientes
            row["_resources"] = dist_by_dataset.get(ds_id, [])

            guid = self._make_guid(ds_id, source_url, i, row)
            content = json.dumps(row, default=str)

            obj = HarvestObject(guid=guid, job=harvest_job, content=content, extras=[])
            obj.save()
            object_ids.append(obj.id)

        log.info("XLSXHarvester: %d datasets recolectados", len(object_ids))
        return object_ids

    # ----------------------------------------------------------- fetch_stage --

    def fetch_stage(self, harvest_object):
        if not harvest_object.content:
            self._save_object_error(
                "El harvest object no tiene contenido", harvest_object, "Fetch"
            )
            return False
        return True

    # ---------------------------------------------------------- import_stage --

    def import_stage(self, harvest_object):
        log.info("XLSXHarvester import_stage: %s", harvest_object.id)

        if not harvest_object.content:
            self._save_object_error("Contenido vacío en import", harvest_object, "Import")
            return False

        self._set_config(harvest_object.source.config)

        try:
            row = json.loads(harvest_object.content)
        except ValueError as e:
            self._save_object_error("JSON inválido: %s" % e, harvest_object, "Import")
            return False

        try:
            package_dict = self._build_package_dict(row, harvest_object)
        except Exception as e:
            self._save_object_error(
                "Error construyendo package_dict: %s" % e, harvest_object, "Import"
            )
            return False

        return self._create_or_update_package(
            package_dict,
            harvest_object,
            package_dict_form="package_show",
        )

    # ---------------------------------------------------------------- helpers --

    def _set_config(self, config_str):
        try:
            self.config = json.loads(config_str) if config_str else {}
        except ValueError:
            self.config = {}

    def _fetch_workbook(self, url):
        ua = self.config.get("user_agent", "ckanext-harvest-xlsx/1.0")
        resp = requests.get(url, headers={"User-Agent": ua}, timeout=60)
        resp.raise_for_status()
        if "html" in resp.headers.get("Content-Type", ""):
            raise ValueError(
                "La URL devolvió HTML en vez de un archivo XLSX. "
                "Verificá que la URL apunte directamente al archivo."
            )
        return openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)

    def _read_sheet(self, wb, sheet_name, column_map, skip=0):
        """
        Lee una hoja del workbook y devuelve una lista de dicts
        mapeados mediante column_map (header_excel → field_name).
        Filas completamente vacías se omiten.
        """
        if sheet_name not in wb.sheetnames:
            log.warning("Hoja '%s' no encontrada en el workbook", sheet_name)
            return []

        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not rows:
            return []

        raw_headers = rows[skip]
        headers = [str(h).strip() if h is not None else "" for h in raw_headers]
        data_rows = rows[skip + 1:]

        result = []
        for row in data_rows:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            raw = {headers[j]: (row[j] if j < len(row) else None) for j in range(len(headers))}
            # Traducir encabezados → field_names; ignorar columnas desconocidas
            mapped = {}
            for col_header, val in raw.items():
                field_name = column_map.get(col_header)
                if field_name and val is not None:
                    mapped[field_name] = str(val).strip()
            result.append(mapped)

        return result

    def _make_guid(self, ds_id, source_url, row_index, row):
        """GUID = URL_fuente/dataset_identifier, o hash como fallback."""
        if ds_id:
            return "%s/%s" % (source_url.rstrip("/"), ds_id)
        title = row.get("dataset_title", "")
        raw = "%s|%s|%d" % (source_url, title, row_index)
        return hashlib.md5(raw.encode("utf-8")).hexdigest()

    def _build_package_dict(self, row, harvest_object):
        """Convierte la fila del dataset (más sus recursos) en un package_dict de CKAN."""
        pkg = {}

        # Campos directos
        field_map = {
            "id" : "id",
            "dataset_title":           "title",
            "name":                    "name",
            "owner_org":               "owner_org",
            "dataset_description":     "notes",
            "dataset_issued":          "metadata_created",
            "dataset_modified":        "metadata_modified",
            "license_id":              "license_id",
            "dataset_publisher_mbox":  "author_email",
        }
        for src, dst in field_map.items():
            val = row.get(src, "").strip()
            if val:
                pkg[dst] = val

        if not pkg.get("title"):
            raise ValueError("El dataset no tiene título (dct:title)")

        pkg["name"] = self._gen_new_name(pkg.get("name") or pkg["title"])

        # Organización por defecto
        if not pkg.get("owner_org"):
            default_org = self.config.get("default_owner_org", "")
            if default_org:
                pkg["owner_org"] = default_org

        # Keywords → tags
        keywords_raw = row.get("dataset_keywords", "")
        if keywords_raw:
            pkg["tags"] = [
                {"name": t.strip()}
                for t in keywords_raw.split(",")
                if t.strip()
            ]

        # Extras: campos DCAT-AP adicionales
        extra_fields = [
            "dataset_status", "dataset_superTheme",
            "dataset_theme", "dataset_accrualPeriodicity",
            "temporal_start", "temporal_end",
            "spatial", "spatial_uri", "spatial_coverage",
            "dataset_source", "dataset_hvdcategory",
        ]
        extras = [{"key": "harvest_source_url", "value": harvest_object.source.url}]
        for f in extra_fields:
            val = row.get(f, "").strip()
            if val:
                extras.append({"key": f, "value": val})
        pkg["extras"] = extras

        # Recursos (distribuciones de la pestaña Distributions)
        resources = []
        for dist in row.get("_resources", []):
            res_url = dist.get("url", "").strip()
            if not res_url:
                continue
            res = {
                "url":    res_url,
                "name":   dist.get("name", pkg["title"]),
                "format": dist.get("format", ""),
                "description": dist.get("description", ""),
            }
            # Campos de distribución adicionales como extras del recurso
            for dist_field in (
                "distribution_identifier", "distribution_mediaType",
                "distribution_type", "distribution_character_set",
                "distribution_scale", "distribution_projection",
                "distribution_iso19115_url", "distribution_wfs_url",
            ):
                val = dist.get(dist_field, "").strip()
                if val:
                    res[dist_field] = val
            resources.append(res)

        # Si no hay distribuciones, apuntar al archivo fuente como recurso genérico
        if not resources:
            resources = [{
                "url":    harvest_object.source.url,
                "format": "XLSX",
                "name":   "Archivo fuente XLSX",
            }]

        pkg["resources"] = resources

        # Buscar si el package ya existe en CKAN (caso update).
        # HarvesterBase vincula harvest_object.package_id al package en runs anteriores,
        # así que si el objeto ya tiene package_id ese es el ID interno correcto.
        # Como fallback buscamos por guid en harvest_object previos del mismo source.
        existing_package_id = self._get_existing_package_id(harvest_object)
        if existing_package_id:
            pkg["id"] = existing_package_id

        # Si no hay package previo no asignamos pkg["id"] y dejamos que CKAN
        # genere un UUID nuevo al crear el package (evita el ForeignKeyViolation).

        return pkg

    def _get_existing_package_id(self, harvest_object):
        """
        Devuelve el ID interno del package en CKAN si el dataset ya fue
        harvested antes, o None si es la primera vez.

        Estrategia:
        1. El harvest_object actual puede ya tener package_id asignado
           (cuando HarvesterBase lo reutiliza de un run anterior).
        2. Si no, buscar en harvest_objects anteriores del mismo source
           que tengan el mismo guid y tengan package_id asignado.
        """
        # Caso 1: el objeto ya trae package_id (run de actualización)
        if harvest_object.package_id:
            return harvest_object.package_id

        # Caso 2: buscar en objetos anteriores del mismo source con el mismo guid
        previous = (
            HarvestObject.Session.query(HarvestObject)
            .filter(HarvestObject.guid == harvest_object.guid)
            .filter(HarvestObject.package_id != None)          # noqa: E711
            .filter(HarvestObject.id != harvest_object.id)
            .order_by(HarvestObject.gathered.desc())
            .first()
        )
        if previous and previous.package_id:
            # Verificar que el package todavía exista en CKAN
            try:
                toolkit.get_action("package_show")(
                    {"ignore_auth": True}, {"id": previous.package_id}
                )
                return previous.package_id
            except toolkit.ObjectNotFound:
                log.warning(
                    "Package %s ya no existe en CKAN, se creará uno nuevo",
                    previous.package_id,
                )

        return None

    def _save_gather_error(self, message, harvest_job):
        log.error("XLSX gather error: %s", message)
        HarvestGatherError(message=message, job=harvest_job).save()
