from __future__ import absolute_import

import io
import json
import logging
import hashlib
import os
import unicodedata
import uuid
import requests
import openpyxl
import datetime

import sqlalchemy as sa

from ckan import model
from ckan.logic import ValidationError, get_action
from ckan.plugins import toolkit
from ckan.views.user import login

from ckanext.harvest.model import HarvestObject, HarvestGatherError
from .base import HarvesterBase

log = logging.getLogger(__name__)


def _normalize_str(s):
    """Minúsculas y sin signos diacríticos para matching de nombres geográficos."""
    s = unicodedata.normalize('NFD', s.lower())
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')
log_ckan = logging.getLogger("ckan.logic.action.create")
log_ckan.setLevel(logging.DEBUG)


# -----------------------------------------------------------------------------
# Mapeo del perfil previo de metadatos versión excel a campos ckan nuevos
# -----------------------------------------------------------------------------
ANDINO_V1_DATASET_COLUMN_MAP = {
    "dataset_identifier": "id",
    "dataset_title": "title",
    "dataset_description": "notes",
    "dataset_publisher_name": "dataset_publisher_name",
    "dataset_publisher_mbox": "dataset_publisher_mbox",
    "dataset_contactPoint_fn": "dataset_contactPoint_fn",
    "dataset_contactPoint_hasEmail": "dataset_contactPoint_hasEmail",
    "dataset_superTheme": "dataset_superTheme",
    "dataset_theme": "dataset_theme",
    "dataset_keyword": "dataset_keywords",
    "dataset_accrualPeriodicity": "dataset_accrualPeriodicity",
    "dataset_issued": "dataset_issued",
    "dataset_modified": "dataset_modified",
    "dataset_language": "dataset_language",
    "dataset_spatial": "spatial",
    "dataset_temporal": "temporal_start",
    "dataset_landingPage": "dataset_landingPage",
    "dataset_license": "license_id",
    "dataset_source": "dataset_source",
}

ANDINO_V1_RESOURCE_COLUMN_MAP = {
    "dataset_identifier": "id",
    "distribution_identifier": "distribution_identifier",
    "distribution_title": "name",
    "distribution_description": "description",
    "distribution_downloadURL": "url",
    "distribution_fileName": "distribution_fileName",
    "distribution_format": "format",
    "distribution_mediaType": "mimetype",
    "distribution_license": "distribution_license",
    "distribution_byteSize": "distribution_byteSize",
    "distribution_modified": "last_modified",
    "distribution_issued": "created",
    "distribution_rights": "distribution_rights",
}

# -----------------------------------------------------------------------------
# Mapeo del perfil nuevo de metadatos versión excel a campos ckan nuevos
# -----------------------------------------------------------------------------
ANDINO_V2_DATASET_COLUMN_MAP = {
    "dataset_identifier": "id",  # PK interna CKAN
    "dataset_title": "title",
    "dataset_publisher_name": "dataset_publisher_name",
    "dataset_publisher_mbox": "dataset_publisher_mbox",
    "dataset_description": "notes",
    "dataset_issued": "dataset_issued",
    "dataset_modified": "dataset_modified",
    "dataset_status": "dataset_status",
    "dataset_theme": "dataset_theme",
    "dataset_superTheme": "dataset_superTheme",
    "dataset_keywords": "dataset_keywords",
    "dataset_hvdCategory": "dataset_hvdCategory",
    "dataset_accrualPeriodicity": "dataset_accrualPeriodicity",
    "temporal_start": "temporal_start",
    "temporal_end": "temporal_end",
    "dataset_spatial": "spatial",
    "spatial_uri": "spatial_uri",
    "dataset_license_id": "license_id",
    "dataset_source": "dataset_source",
}

ANDINO_V2_RESOURCE_COLUMN_MAP = {
    "dataset_identifier": "id",  # FK → dataset
    "distribution_identifier": "distribution_identifier",  # identificador de la distribución
    "distribution_download_url": "url",  # dcat:downloadURL
    "distribution_name": "name",  # dct:title del recurso ← "name" en v2
    "distribution_description": "description",  # dct:description
    "distribution_format": "format",  # dct:format
    "distribution_mediaType": "mediaType",  # dcat:mediaType #
    "distribution_character_set": "character_set",  # cnt:characterEncoding
    "distribution_scale": "scale",  # escala geográfica
    "distribution_projection": "projection",  # dct:conformsTo (proyección)
    "distribution_iso19115_url": "iso19115_url",  # URL metadato ISO 19115
    "distribution_wfs_url": "wfs_url",  # URL servicio WFS
}

DEFAULT_DATASET_SHEET      = ["dataset", "Dataset", "datasets", "Datasets"]
DEFAULT_DISTRIBUTION_SHEET = ["distribution", "Distribution", "distributions", "Distributions"]


def _resolve_sheet_name(wb, candidates):
    """Retorna el primer nombre de hoja del workbook que coincida con alguno de los candidatos."""
    if isinstance(candidates, str):
        candidates = [candidates]
    return next((s for s in wb.sheetnames if s in candidates), None)

DATASET_REQUIRED      = {"title"}
DISTRIBUTION_REQUIRED = {"url"}

# Columnas del mapping V1 que pueden no estar en el Excel (opcionales/recomendadas)
# La validación de headers será permisiva para estas.
ANDINO_V1_DATASET_OPTIONAL = {
    "dataset_publisher_mbox",
    "dataset_contactPoint_fn",
    "dataset_contactPoint_hasEmail",
    "dataset_theme",
    "dataset_keyword",
    "dataset_modified",
    "dataset_language",
    "dataset_spatial",
    "dataset_temporal",
    "dataset_landingPage",
    "dataset_license",
}
ANDINO_V1_RESOURCE_OPTIONAL = {
    "distribution_description",
    "distribution_fileName",
    "distribution_format",
    "distribution_mediaType",
    "distribution_license",
    "distribution_byteSize",
    "distribution_modified",
    "distribution_rights",
}


class XLSXHarvester(HarvesterBase):
    """
    Harvester para archivos Excel (.xlsx) con el perfil de metadatos datos.gob.ar.

    Implementa las tres etapas del ciclo de harvest de CKAN (gather, fetch, import)
    para fuentes XLSX. Soporta dos versiones del perfil de metadatos:
      - V2 (preferido): columnas con nombres del perfil actual (dataset_description,
        distribution_download_url, etc.)
      - V1 (fallback): columnas del perfil legacy de Andino

    Estructura esperada del archivo:
      - Pestaña "dataset"      : una fila por dataset; encabezados según el perfil activo
      - Pestaña "distribution" : una fila por distribución; incluye dataset_identifier (FK)

    El harvester intenta primero el mapeo V2 y cae a V1 si los headers no coinciden.

    Configuración opcional (JSON en el campo "Configuration" del formulario):
    {
        "dataset_sheet":      "dataset",
        "distribution_sheet": "distribution",
        "skip_rows":          0,
        "default_owner_org":  "mi-org",
        "default_tags":       ["tag1", "tag2"],
        "default_extras":     {"clave": "valor"},
        "override_extras":    false,
        "user_agent":         "mi-portal/1.0",
        "force_all":          false
    }

    Nota: default_tags es una lista de strings (nombres de tags), no de dicts.
    """

    config = None

    # ------------------------------------------------------------------ info --

    def info(self):
        """Retorna los metadatos de registro del harvester para CKAN (nombre, título, descripción)."""
        return {
            "name": "xlsx_harvester",
            "title": "XLSX",
            "description": (
                "Harvests remote XLSX files following the datos.gob.ar DCAT-AP schema. "
                "Applies the same field transformations as the CKAN harvester. "
                "Expects a 'dataset' sheet and a 'distribution' sheet."
            ),
            "form_config_interface": "Text",
        }

    # --------------------------------------------------------- validate_config --

    def validate_config(self, config_str):
        """
        Valida el JSON de configuración del harvest source.

        Verifica tipos de cada clave conocida: dataset_sheet, distribution_sheet,
        default_owner_org y user_agent deben ser strings; skip_rows debe ser entero;
        default_tags debe ser lista de dicts; default_extras debe ser dict; y
        read_only, force_all, override_extras deben ser booleanos.
        Lanza ValueError ante cualquier violación de tipo o JSON inválido.
        Retorna config_str sin modificar si todo es válido.
        """
        if not config_str:
            return config_str
        try:
            cfg = json.loads(config_str)
        except ValueError as e:
            raise ValueError("La configuración debe ser un JSON válido: %s" % e)

        for key in ("dataset_sheet", "distribution_sheet", "default_owner_org", "user_agent"):
            if key in cfg and not isinstance(cfg[key], str):
                raise ValueError("%s debe ser un string" % key)
        if "skip_rows" in cfg and not isinstance(cfg["skip_rows"], int):
            raise ValueError("skip_rows debe ser un entero")
        if "default_tags" in cfg:
            if not isinstance(cfg["default_tags"], list):
                raise ValueError("default_tags debe ser una lista")
            if cfg["default_tags"] and not isinstance(cfg["default_tags"][0], dict):
                raise ValueError("default_tags debe ser una lista de diccionarios")
        if "default_extras" in cfg and not isinstance(cfg["default_extras"], dict):
            raise ValueError("default_extras debe ser un diccionario")
        for key in ("read_only", "force_all", "override_extras"):
            if key in cfg and not isinstance(cfg[key], bool):
                raise ValueError("%s debe ser booleano" % key)

        return config_str

    # ------------------------------------------------------ get_original_url --

    def get_original_url(self, harvest_object_id):
        """Retorna la URL del harvest source asociado al harvest_object_id dado, o None si no existe."""
        obj = HarvestObject.get(harvest_object_id)
        return obj.source.url if obj else None

    # ------------------------------------------------------- fields_options --

    @property
    def fields_options(self):
        """
        Carga y cachea el archivo assets/fields_options.json.

        El JSON contiene los diccionarios de traducción de valores controlados:
        superthemes, province_codes, mimetypes y category_formats.
        Se lee una sola vez y se guarda en _field_options. Retorna {} si el
        archivo no existe o no puede parsearse.
        """
        if not hasattr(self, "_field_options"):
            self._field_options = None
        if self._field_options:
            return self._field_options
        path = os.path.join(os.path.dirname(__file__), "../assets/fields_options.json")
        try:
            with open(path, "r") as f:
                self._field_options = json.load(f)
        except Exception as e:
            log.warning("Error cargando fields_options.json en %s: %s", path, e)
            self._field_options = {}
        return self._field_options

    def get_field_options(self, field_name):
        """Retorna el dict de opciones para field_name desde fields_options.json, o {} si no existe."""
        return self.fields_options.get(field_name, {})

    # ---------------------------------------------------------- gather_stage --
    def _validate_mapping_headers(self, workbook, sheet_name, mapping, skip=0, optional_keys=None):
        """
        Valida que los headers requeridos del mapping estén presentes en la hoja Excel.

        Lee la fila de encabezados (respetando skip), normaliza a lowercase y
        compara contra las claves del mapping. Las claves en optional_keys se excluyen
        de la verificación. Lanza ValueError si la hoja no existe o si faltan columnas
        requeridas, listando cuáles son.
        """
        optional_keys = {k.lower() for k in (optional_keys or set())}

        if sheet_name is None:
            raise ValueError("No existe la hoja '%s'" % sheet_name)
        try:
            ws = workbook[sheet_name]
        except KeyError:
            raise ValueError("No existe la hoja '%s'" % sheet_name)

        header_row = next(
            ws.iter_rows(
                min_row=skip + 1,
                max_row=skip + 1,
                values_only=True,
            )
        )

        excel_headers = {
            str(h).strip().lower()
            for h in header_row
            if h is not None
        }

        required_headers = {
            str(h).strip().lower()
            for h in mapping.keys()
            if str(h).strip().lower() not in optional_keys
        }

        missing = sorted(required_headers - excel_headers)

        if missing:
            raise ValueError(
                "Faltan columnas requeridas en hoja '%s': %s"
                % (sheet_name, ", ".join(missing))
            )

    def gather_stage(self, harvest_job):
        """
        Descarga el XLSX, lo parsea y crea un HarvestObject por cada dataset.

        Flujo:
        1. Descarga el archivo desde harvest_job.source.url.
        2. Intenta leer con el mapeo V2 y, si falla, con V1; aborta si ninguno es válido.
        3. Agrupa las filas de distribuciones por dataset_identifier (FK).
        4. Por cada fila de dataset: valida campos requeridos, genera un GUID estable
           (url_fuente/dataset_identifier o MD5 como fallback), descarta duplicados,
           adjunta la lista de distribuciones en _resources y persiste el HarvestObject
           con el contenido serializado como JSON.
        5. Marca como no-current los GUIDs del run anterior que ya no aparecen en el Excel.
        Retorna la lista de IDs de HarvestObjects creados.
        """
        log.info("XLSXHarvester gather_stage: %s", harvest_job.source.url)

        self._set_config(harvest_job.source.config)
        source_url = harvest_job.source.url.strip()

        try:
            wb = self._fetch_workbook(source_url)
        except Exception as e:
            self._save_gather_error(
                "No se pudo descargar/leer el XLSX: %s" % e,
                harvest_job,
            )
            return []

        skip       = self.config.get("skip_rows", 0)
        ds_sheet   = _resolve_sheet_name(wb, self.config.get("dataset_sheet", DEFAULT_DATASET_SHEET))
        dist_sheet = _resolve_sheet_name(wb, self.config.get("distribution_sheet", DEFAULT_DISTRIBUTION_SHEET))

        ds_rows = dist_rows = None
        errors = []

        for ds_map, dist_map, version, ds_optional, dist_optional in [
            (
                ANDINO_V2_DATASET_COLUMN_MAP,
                ANDINO_V2_RESOURCE_COLUMN_MAP,
                "V2",
                set(),   # V2: validación estricta, sin opcionales
                set(),
            ),
            (
                ANDINO_V1_DATASET_COLUMN_MAP,
                ANDINO_V1_RESOURCE_COLUMN_MAP,
                "V1",
                ANDINO_V1_DATASET_OPTIONAL,   # V1: columnas que pueden faltar
                ANDINO_V1_RESOURCE_OPTIONAL,
            ),
        ]:
            try:
                self._validate_mapping_headers(wb, ds_sheet,   ds_map,   skip, ds_optional)
                self._validate_mapping_headers(wb, dist_sheet, dist_map, skip, dist_optional)

                ds_rows_try   = self._read_sheet(wb, ds_sheet,   ds_map,   skip)
                dist_rows_try = self._read_sheet(wb, dist_sheet, dist_map, skip)

            except Exception as e:
                errors.append("Mapeo %s inválido: %s" % (version, e))
                continue

            ds_rows   = ds_rows_try
            dist_rows = dist_rows_try
            log.info("Catálogo leído correctamente con mapeo %s", version)
            break

        if ds_rows is None:
            self._save_gather_error(
                "No se pudo leer el XLSX con ningún mapeo conocido. Detalles: %s"
                % " | ".join(errors),
                harvest_job,
            )
            return []

        # Agrupar distribuciones por dataset_identifier (FK)
        dist_by_dataset = {}
        for d in dist_rows:
            ds_id = str(d.get("id", "")).strip()
            if ds_id:
                dist_by_dataset.setdefault(ds_id, []).append(d)

        # GUIDs de datasets ya harvested en runs anteriores (para detectar eliminados)
        previous_guids = {
            r[0]
            for r in model.Session.query(HarvestObject.guid).filter(
                HarvestObject.harvest_source_id == harvest_job.source.id,
                HarvestObject.current == True,  # noqa: E712
            ).all()
        }

        object_ids = []
        seen_guids = set()

        for i, row in enumerate(ds_rows):
            ds_id = str(row.get("id", "")).strip()

            missing = [
                f for f in DATASET_REQUIRED
                if not str(row.get(f, "")).strip()
            ]
            if missing:
                log.warning("Fila %d omitida: faltan campos %s", i + 2, missing)
                continue

            guid = self._make_guid(ds_id, source_url, i, row)

            if guid in seen_guids:
                log.info("Dataset duplicado descartado: %s", guid)
                continue

            seen_guids.add(guid)
            row["_resources"] = dist_by_dataset.get(ds_id, [])

            content = json.dumps(row, default=str, sort_keys=True)

            obj = HarvestObject(
                guid=guid,
                job=harvest_job,
                content=content,
                extras=[],
            )
            obj.save()
            object_ids.append(obj.id)

        # Marcar como no-current los datasets que ya no están en el Excel
        deleted_guids = previous_guids - seen_guids
        if deleted_guids:
            log.info(
                "XLSXHarvester: %d datasets ya no están en el Excel y serán marcados como no-current",
                len(deleted_guids),
            )
            model.Session.query(HarvestObject).filter(
                HarvestObject.guid.in_(deleted_guids),
                HarvestObject.harvest_source_id == harvest_job.source.id,
            ).update({"current": False}, synchronize_session=False)
            model.Session.commit()

        log.info("XLSXHarvester: %d datasets recolectados", len(object_ids))
        return object_ids

    # ----------------------------------------------------------- fetch_stage --

    def fetch_stage(self, harvest_object):
        """
        Verifica que el HarvestObject tenga contenido serializado.

        El contenido ya fue embebido como JSON durante gather_stage, por lo que
        no se realiza ninguna descarga. Retorna True si content está presente,
        False (con error registrado) si está vacío.
        """
        if not harvest_object.content:
            self._save_object_error(
                "El harvest object no tiene contenido", harvest_object, "Fetch"
            )
            return False
        return True

    # ---------------------------------------------------------- import_stage --

    def import_stage(self, harvest_object):
        """
        Persiste el dataset en CKAN a partir del contenido del HarvestObject.

        Flujo:
        1. Deserializa el JSON almacenado en harvest_object.content.
        2. Llama a _build_package_dict para construir los campos core de CKAN
           (id, name, owner_org, resources, etc.).
        3. Llama a modify_package_dict para aplicar transformaciones de dominio
           (superTheme, accrualPeriodicity, temporal, spatial, etc.).
        4. Inyecta los default_tags de la configuración si no están ya presentes.
        5. Si el package ya existe y no está borrado, ejecuta package_update
           marcando primero el HarvestObject previo como no-current; si no existe,
           ejecuta package_create difiriendo la FK harvest_object_package_id_fkey
           para evitar conflictos de orden de inserción.
        Retorna True si la operación fue exitosa, lanza la excepción en caso contrario.
        """
        log.debug("XLSXHarvester import_stage: %s", harvest_object.id)

        if not harvest_object.content:
            self._save_object_error(
                "Contenido vacío en el objeto de harvest", harvest_object, "Import"
            )
            return False

        self._set_config(harvest_object.source.config)


        # 1. Deserializar la fila guardada en el gather_stage
        try:
            row = json.loads(harvest_object.content)
        except ValueError as e:
            self._save_object_error(
                "JSON inválido en content: %s" % e, harvest_object, "Import"
            )
            return False

        # 2. Construir el package_dict base (campos core de CKAN)
        try:
            package_dict = self._build_package_dict(row, harvest_object)
        except Exception as e:
            self._save_object_error(
                "Error construyendo package_dict: %s" % e, harvest_object, "Import"
            )
            return False

        # 3. Aplicar transformaciones específicas (Theme, SuperTheme, campos temporales, etc.)
        try:
            package_dict = self.modify_package_dict(package_dict, harvest_object)


        except Exception as e:
            self._save_object_error(
                "Error en modify_package_dict: %s" % e, harvest_object, "Import"
            )
            return False

        # 4. Inyectar tags por defecto desde la configuración del harvest source
        # default_tags es una lista de strings (nombres de tags)
        default_tags = self.config.get("default_tags", [])
        if default_tags:
            package_dict.setdefault("tags", [])
            existing_tag_names = {t["name"] for t in package_dict["tags"] if "name" in t}
            for tag_name in default_tags:
                if isinstance(tag_name, str) and tag_name not in existing_tag_names:
                    package_dict["tags"].append({"name": tag_name})

        # 5. Persistir en CKAN
        context = {
            "model": model,
            "session": model.Session,
            "user": self._get_user_name(),
            "ignore_auth": True,
        }

        try:
            exists = model.Package.get(package_dict["id"])

            if exists and exists.state != "deleted":
                # Comparar contenido con el HarvestObject previo para evitar
                # actualizar si los metadatos no cambiaron
                prev_object = (
                    model.Session.query(HarvestObject)
                    .filter(
                        HarvestObject.guid == harvest_object.guid,
                        HarvestObject.current == True,  # noqa: E712
                        HarvestObject.id != harvest_object.id,
                    )
                    .first()
                )
                if prev_object and prev_object.content == harvest_object.content:
                    log.info(
                        "Sin cambios en dataset %s, saltando actualización",
                        package_dict["name"],
                    )
                    harvest_object.package_id = package_dict["id"]
                    harvest_object.current = True
                    harvest_object.report_status = 'not modified'
                    harvest_object.add()
                    model.Session.query(HarvestObject).filter(
                        HarvestObject.package_id == package_dict["id"],
                        HarvestObject.id != harvest_object.id,
                    ).update({"current": False})
                    model.Session.commit()
                    return 'unchanged'

                log.info("Actualizando dataset existente: %s", package_dict["name"])

                # Marcar como no-current cualquier HarvestObject previo de este paquete
                model.Session.query(HarvestObject).filter(
                    HarvestObject.package_id == package_dict["id"]
                ).update({"current": False})

                # Setear este HarvestObject como current ANTES del update para
                # que el hook before_dataset_index encuentre el harvest_source_id
                harvest_object.package_id = package_dict["id"]
                harvest_object.current = True
                harvest_object.add()
                model.Session.flush()

                result = get_action("package_update")(context, package_dict)
            else:
                log.info("Creando nuevo dataset: %s", package_dict["name"])
                harvest_object.package_id = package_dict['id']
                harvest_object.current = True
                harvest_object.add()

                # Diferir el FK porque el package todavía no existe en BD
                model.Session.execute(
                    sa.text("SET CONSTRAINTS harvest_object_package_id_fkey DEFERRED")
                )
                model.Session.flush()

                result = get_action("package_create")(context, package_dict)

            model.Session.commit()
            return True

        except sa.exc.IntegrityError as e:
            model.Session.rollback()
            orig = getattr(e, "orig", None)
            orig_type = type(orig).__name__ if orig else ""
            if orig_type == "UniqueViolation" and "resource_pkey" in str(e):
                resources = package_dict.get("resources", [])
                ids = [r.get("id") for r in resources if r.get("id")]
                dupes = sorted({i for i in ids if ids.count(i) > 1})
                self._save_object_error(
                    "IDs de distribución duplicados en el dataset '%s': %s. "
                    "Revisá que los distribution_identifier sean únicos en la fuente."
                    % (package_dict.get("name", "?"), ", ".join(dupes)),
                    harvest_object,
                    "Import",
                )
            else:
                self._save_object_error(
                    "Error de integridad al guardar en CKAN: %s" % str(e),
                    harvest_object,
                    "Import",
                )
        except Exception as e:
            model.Session.rollback()
            self._save_object_error(
                "Error al guardar en CKAN (Acción API): %s" % str(e),
                harvest_object,
                "Import",
            )
            raise e

    # ---------------------------------------------------- modify_package_dict --

    def modify_package_dict(self, package_dict, harvest_object):
        """
        Aplica transformaciones de dominio al package_dict para alinearlo con el perfil V2.

        Transformaciones en orden:
        - dataset_status: asigna "Completed" (ADMS) si está vacío.
        - dataset_hvdCategory: asigna "no aplica" si está vacío.
        - spatial: si el valor no es GeoJSON válido, intenta interpretar los candidatos como
          códigos de provincia y los traduce a URIs vía province_codes; deja spatial_uri con
          la lista de URIs y vacía spatial. Si no hay coincidencias, vacía ambos campos.
        - temporal_start: normaliza el sufijo "Z" a "+00:00". Si contiene "/" (intervalo ISO 8601),
          lo parte en temporal_start y temporal_end. Si no hay temporal_end, lo iguala a temporal_start.
        - dataset_accrualPeriodicity: se preserva el valor tal como viene, sin mapeo.
        - dataset_superTheme: parsea string CSV o array JSON, mapea siglas a URLs vía superthemes;
          asigna ["Sin tema"] si el resultado es vacío.
        - dataset_theme: sobreescribe siempre con "Tema específico 1".
        - dataset_issued / dataset_modified: si están vacíos, los toma de metadata_created /
          metadata_modified del paquete existente en CKAN.
        - dataset_language: asigna siempre la URI del español (EU Publications Office).
        - resources: por cada distribución, rellena con "" los campos opcionales faltantes
          (character_set, scale, projection, iso19115_url, wfs_url) y calcula category
          mapeando el format contra category_formats.
        """

        # 2. dataset_status → solo se agrega si falta
        if not package_dict.get("dataset_status"):
            title = package_dict.get("title", "")
            if "discontinuado" in title.lower():
                package_dict["dataset_status"] = "http://purl.org/adms/status/Withdrawn"
            else:
                package_dict["dataset_status"] = "http://purl.org/adms/status/Completed"

        # 3. dataset_hvdCategory → solo se agrega si falta
        if not package_dict.get("dataset_hvdCategory"):
            package_dict["dataset_hvdCategory"] = "no aplica"

        # Revisa campo spatial para ver si puede guardar algún valor. Asume que si ya
        # es geojson no hay necesidad de hacer ninguna operacion con spatial_uri
        spatial = package_dict.get("spatial", "")
        is_geojson = False
        if spatial:
            try:
                parsed = json.loads(spatial)
                if isinstance(parsed, dict) and parsed.get("type"):
                    is_geojson = True
            except (ValueError, TypeError):
                pass

        if not is_geojson:
            if isinstance(spatial, list):
                candidates = [str(v).strip() for v in spatial if str(v).strip()]
            elif spatial:
                try:
                    parsed = json.loads(spatial)
                    candidates = parsed if isinstance(parsed, list) else [str(parsed).strip()]
                except (ValueError, TypeError):
                    candidates = [v.strip() for v in str(spatial).split(",") if v.strip()]
            else:
                candidates = []

            province_codes = self.get_field_options("province_codes")
            matched_uris = [province_codes[c] for c in candidates if c in province_codes]

            if matched_uris:
                package_dict["spatial_uri"] = matched_uris
                package_dict["spatial"] = ""
            else:
                package_dict["spatial_uri"] = ""
                package_dict["spatial"] = ""

        # Fallback: buscar nombre de provincia/región en el título si spatial_uri sigue vacío
        if not package_dict.get("spatial_uri"):
            title_norm = _normalize_str(package_dict.get("title", ""))
            province_names = self.get_field_options("province_names")
            matched_uris = list({
                uri for name, uri in province_names.items()
                if name in title_norm
            })
            if matched_uris:
                package_dict["spatial_uri"] = matched_uris

        # 5. temporal_end → se deriva de temporal_start si contiene "/" (formato ISO 8601 interval)
        temporal_start = str(package_dict.get("temporal_start", "") or "")
        temporal_start = temporal_start.replace("Z", "+00:00")

        if not temporal_start:
            pass
        elif "/" in temporal_start:
            parts = temporal_start.split("/", 1)
            package_dict['temporal_start'] = parts[0]
            package_dict['temporal_end'] = parts[1]
        else:
            package_dict['temporal_start'] = temporal_start
            if not package_dict.get('temporal_end'):
                package_dict['temporal_end'] = temporal_start

        # 6. dataset_accrualPeriodicity: se preserva el valor que venga; default si vacío
        if not package_dict.get("dataset_accrualPeriodicity"):
            package_dict["dataset_accrualPeriodicity"] = "sin especificar"

        # 7. dataset_superTheme → sustituir siglas por URLs; default si vacío
        DEFAULT_SUPER_THEME = "Sin tema"
        raw_st = package_dict.get("dataset_superTheme")

        if isinstance(raw_st, str):
            try:
                st_list = json.loads(raw_st)
            except (ValueError, TypeError):
                st_list = [v.strip() for v in raw_st.split(",") if v.strip()]
        elif isinstance(raw_st, list):
            st_list = raw_st
        else:
            st_list = []

        if st_list:
            superthemes = self.get_field_options("superthemes")
            package_dict["dataset_superTheme"] = [
                superthemes.get(item, item) for item in st_list
            ]
        else:
            package_dict["dataset_superTheme"] = [DEFAULT_SUPER_THEME]

        # 8. dataset_theme → siempre se sobrescribe
        package_dict["dataset_theme"] = "Tema específico 1"

        package_dict['dataset_language'] = "http://publications.europa.eu/resource/authority/language/SPA"

        # 10. Campos faltantes en recursos/distribuciones
        DISTRIBUTION_DEFAULTS = {
            "character_set": "",
            "scale": "",
            "projection": "",
            "iso19115_url": "",
            "wfs_url": ""
        }
        for resource in package_dict.get("resources", []):
            for field, default in DISTRIBUTION_DEFAULTS.items():
                if not resource.get(field):
                    resource[field] = default

            # Resolver mimetype → URI IANA solo si no es ya una URI
            media_type = resource.get("mediaType", "") or resource.get("mimetype", "")
            if not media_type:
                fmt = resource.get("format", "").lower()
                media_type = self.get_field_options("format_to_mediatype").get(fmt, "")
            if media_type.startswith("http"):
                resource["mimetype"] = media_type
            else:
                resource["mimetype"] = (
                    self.get_field_options("mimetypes").get(media_type) or "other"
                )

            # Resolver category según formato solo si no viene ya como URI
            existing_category = resource.get("category", "")
            if existing_category and existing_category.startswith("http"):
                pass
            else:
                resource_format = resource.get("format", "").lower()
                resource["category"] = next(
                    (k for k, v in self.get_field_options("category_formats").items()
                     if resource_format in v),
                    "other"
                )

        return package_dict

    # ---------------------------------------------------------------- helpers --

    def _set_config(self, config_str):
        """Parsea config_str como JSON y lo guarda en self.config. Asigna {} si está vacío o es inválido."""
        try:
            self.config = json.loads(config_str) if config_str else {}
        except ValueError:
            self.config = {}

    def _fetch_workbook(self, url):
        """
        Descarga el archivo XLSX desde url y retorna un openpyxl Workbook.

        Usa el User-Agent configurado (o "ckanext-harvest-xlsx/1.0" por defecto).
        Lanza ValueError si la respuesta es HTML (indicando que la URL no apunta
        directamente al archivo). Abre el workbook en modo read-only y data-only
        para mayor eficiencia con archivos grandes.
        """
        ua = self.config.get("user_agent", "ckanext-harvest-xlsx/1.0")
        resp = requests.get(url, headers={"User-Agent": ua}, timeout=60)
        resp.raise_for_status()
        if "html" in resp.headers.get("Content-Type", ""):
            raise ValueError(
                "La URL devolvió HTML en vez de un archivo XLSX. "
                "Verificá que la URL apunte directamente al archivo."
            )
        return openpyxl.load_workbook(
            io.BytesIO(resp.content), read_only=True, data_only=True
        )

    def _read_sheet(self, wb, sheet_name, column_map, skip=0):
        """
        Lee una hoja del workbook y retorna una lista de dicts con campos internos.

        Omite las primeras `skip` filas, toma la siguiente como fila de encabezados
        y mapea cada columna al nombre interno definido en column_map. Descarta filas
        completamente vacías y columnas que no estén en column_map. Todos los valores
        se convierten a strings y se les aplica strip().
        """
        if sheet_name not in wb.sheetnames:
            log.warning("Hoja '%s' no encontrada en el workbook", sheet_name)
            return []

        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not rows:
            return []

        raw_headers = rows[skip]
        headers = [str(h).strip() if h is not None else "" for h in raw_headers]

        result = []
        for row in rows[skip + 1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            raw = {
                headers[j]: (row[j] if j < len(row) else None)
                for j in range(len(headers))
            }
            mapped = {}
            for col_header, val in raw.items():
                field_name = column_map.get(col_header)
                if field_name:
                    mapped[field_name] = str(val).strip() if val is not None else ""
            result.append(mapped)

        return result

    def _make_guid(self, ds_id, source_url, row_index, row):
        """
        Genera un GUID estable para identificar un dataset entre runs.

        Si ds_id está presente, retorna "source_url/ds_id".
        Si no, calcula el MD5 de "source_url|title|row_index" como fallback,
        usando el título del dataset o vacío si tampoco está disponible.
        """
        if ds_id:
            return "%s/%s" % (source_url.rstrip("/"), ds_id)
        title = row.get("title", row.get("dataset_title", ""))
        raw   = "%s|%s|%d" % (source_url, title, row_index)
        return hashlib.md5(raw.encode("utf-8")).hexdigest()

    def _build_package_dict(self, row, harvest_object):
        """
        Construye el package_dict base con los campos core requeridos por CKAN.

        Transformaciones:
        - Lanza ValueError si el dataset no tiene título.
        - name: genera un slug único con _gen_new_name a partir del name o title.
        - owner_org: si no viene en la fila, lo toma de default_owner_org en la config
          o del owner_org del harvest source.
        - harvest_source_url: agrega la URL fuente para trazabilidad.
        - id del dataset: si hay raw_id en la fila, genera un UUID5 determinístico con
          semilla "owner_org/raw_id" (mismo input → mismo UUID en cada corrida); si no
          hay raw_id, genera un UUID4. Guarda el raw_id original en original_identifier.
        - resources: por cada distribución en _resources (ya con claves target traducidas
          por _read_sheet), construye un dict con url, name, format, description más
          cualquier campo extra que no sea core; omite las que no tienen url. El ID se
          genera con UUID5 ("pkg_id/distribution_identifier") si hay identificador, o
          UUID4 si no. Guarda el identificador original en original_identifier si difiere.
        - Si no hay distribuciones válidas, agrega un recurso de auxilio apuntando
          al archivo XLSX fuente.
        """
        pkg = dict(row)

        if not pkg.get("title"):
            raise ValueError("El dataset no tiene título válido para CKAN")

        pkg["name"] = self._gen_new_name(pkg.get("name") or pkg["title"])

        # Organización dueña
        if not pkg.get("owner_org"):
            default_org = self.config.get("default_owner_org", "")
            if not default_org:
                source_dataset = toolkit.get_action("package_show")(
                    {"ignore_auth": True}, {"id": harvest_object.source.id}
                )
                default_org = source_dataset.get("owner_org", "")
            if default_org:
                pkg["owner_org"] = default_org

        # Keywords → Tags
        #if "dataset_keywords" in pkg and not pkg.get("tags"):
        #    keywords_raw = pkg.pop("dataset_keywords") or ""
        #    if keywords_raw:
        #        pkg["tags"] = [
        #            {"name": t.strip()}
        #            for t in keywords_raw.split(",")
        #            if t.strip()
        #        ]

        # Trazabilidad del harvest
        pkg["harvest_source_url"] = harvest_object.source.url

        # ID único del dataset (se resuelve antes de los recursos para poder
        # usarlo como semilla en el UUID5 de cada distribución)
        raw_id = pkg.get("id", "").strip()
        if raw_id:
            text = "%s/%s" % (pkg.get("owner_org", ""), raw_id)
            pkg["id"] = str(uuid.uuid5(uuid.NAMESPACE_DNS, text))
        else:
            pkg["id"] = str(uuid.uuid4())

        if raw_id and raw_id != pkg["id"]:
            pkg["original_identifier"] = raw_id

        # Procesar distribuciones (_resources → resources)
        resources_raw  = pkg.pop("_resources", [])
        resources_list = []

        CORE_DIST_KEYS = {"url", "name", "format", "description", "id", "distribution_identifier"}

        for dist in resources_raw:
            res_url = dist.get("url", "").strip()
            if not res_url:
                continue

            res = {
                "url":         res_url,
                "name":        dist.get("name") or pkg["title"],
                "format":      (dist.get("format") or "").strip().upper(),
                "description": dist.get("description", ""),
            }

            # Campos extendidos: todo lo que no es un campo core ya mapeado
            for clave, valor in dist.items():
                if clave not in CORE_DIST_KEYS:
                    res[clave] = str(valor or "").strip()

            # ID único de la distribución (misma lógica que el dataset)
            raw_dist_id = dist.get("distribution_identifier", "").strip()
            if raw_dist_id:
                text = "%s/%s" % (pkg["id"], raw_dist_id)
                res["id"] = str(uuid.uuid5(uuid.NAMESPACE_DNS, text))
            else:
                res["id"] = str(uuid.uuid4())

            if raw_dist_id and raw_dist_id != res["id"]:
                res["original_identifier"] = raw_dist_id


            resources_list.append(res)

        # Recurso de auxilio si no hay distribuciones válidas
        if not resources_list:
            resources_list = [{
                "url":    harvest_object.source.url,
                "format": "XLSX",
                "name":   "Archivo fuente XLSX",
            }]

        pkg["resources"] = resources_list

        return pkg

    def _format_super_theme(self, raw_value):
        """
        Normaliza raw_value al formato de lista JSON que espera modify_package_dict.

        Acepta una sigla simple ("AGRI"), una lista separada por comas ("AGRI, SOCI")
        o un array JSON ya serializado ('["AGRI"]'). Retorna "" para valores vacíos.
        """
        if not raw_value:
            return ""
        raw_value = raw_value.strip()
        if raw_value.startswith("["):
            return raw_value
        items = [v.strip() for v in raw_value.split(",") if v.strip()]
        return json.dumps(items)

    def _save_gather_error(self, message, harvest_job):
        """Loguea el mensaje como error y persiste un HarvestGatherError asociado al harvest_job."""
        log.error("XLSXHarvester gather error: %s", message)
        HarvestGatherError(message=message, job=harvest_job).save()
