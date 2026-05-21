from __future__ import absolute_import

import io
import json
import logging
import hashlib
import os
import uuid
import requests
import openpyxl
import datetime

from ckan import model
from ckan.logic import ValidationError, get_action
from ckan.plugins import toolkit

from ckanext.harvest.model import HarvestObject, HarvestGatherError
from .base import HarvesterBase

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Mapeo de encabezados Excel → nombre de campo interno
# ---------------------------------------------------------------------------

# -----------------------------------------------------------------------------
# Mapeo del perfil previo de metadatos versión excel a campos ckan nuevos
# -----------------------------------------------------------------------------
ANDINO_V1_DATASET_COLUMN_MAP = {
    "dataset_identifier": "id",  # PK interna CKAN
    "dataset_title": "title",
    "dataset_description": "dataset_description",
    "dataset_publisher_name": "dataset_publisher_name",
    "dataset_publisher_mbox": "dataset_publisher_mbox",
    "dataset_superTheme": "dataset_superTheme",
    "dataset_theme": "dataset_theme",
    "dataset_keyword": "dataset_keywords",
    "dataset_accrualPeriodicity": "dataset_accrualPeriodicity",
    "dataset_issued": "dataset_issued",
    "dataset_modified": "dataset_modified",
    "dataset_language": "dataset_language",
    "dataset_spatial": "spatial_uri",
    "dataset_temporal": "temporal_start",
    "dataset_landingPage": "dataset_source",
    "dataset_license": "license_id",
    "dataset_source": "dataset_source",
}

ANDINO_V1_RESOURCE_COLUMN_MAP = {
    "dataset_identifier": "id",
    "distribution_identifier": "distribution_identifier",
    "distribution_title": "distribution_name",
    "distribution_description": "distribution_description",
    "distribution_downloadURL": "distribution_download_url",
    "distribution_format": "distribution_format",
    "distribution_mediaType": "distribution_mediaType",
    "distribution_type": "distribution_type",
}

# -----------------------------------------------------------------------------
# Mapeo del perfil nuevo de metadatos versión excel a campos ckan nuevos
# -----------------------------------------------------------------------------
ANDINO_V2_DATASET_COLUMN_MAP = {
    "dataset_identifier": "id",  # PK interna CKAN
    "dataset_title": "title",
    "dataset_publisher_name": "dataset_publisher_name",
    "dataset_publisher_mbox": "dataset_publisher_mbox",
    "dataset_description": "dataset_description",
    "dataset_issued": "dataset_issued",
    "dataset_modified": "dataset_modified",
    "dataset_status": "dataset_status",
    "dataset_theme": "dataset_theme",
    "dataset_superTheme": "dataset_superTheme",
    "dataset_keywords": "dataset_keywords",
    "dataset_hvdCategory": "dataset_hvdCategory",
    "dataset_accrualPeriodicity": "dataset_accrualPeriodicity",
    "temporal_start": "temporal_start",
    "temporal_end": "temporal_end",
    "dataset_spatial": "spatial",
    "spatial_uri": "spatial_uri",
    "dataset_license_id": "license_id",
    "dataset_source": "dataset_source",
}

ANDINO_V2_RESOURCE_COLUMN_MAP = {
    "dataset_identifier": "id",  # FK → dataset
    "distribution_identifier": "distribution_identifier",  # identificador de la distribución
    "distribution_download_url": "distribution_download_url",  # dcat:downloadURL
    "distribution_name": "name",  # dct:title del recurso ← "name" en v2
    "distribution_description": "distribution_description",  # dct:description
    "distribution_format": "distribution_format",  # dct:format
    "distribution_mediaType": "distribution_mediaType",  # dcat:mediaType
    "distribution_type": "distribution_type",  # dct:type
    "distribution_character_set": "distribution_character_set",  # cnt:characterEncoding
    "distribution_scale": "distribution_scale",  # escala geográfica
    "distribution_projection": "distribution_projection",  # dct:conformsTo (proyección)
    "distribution_iso19115_url": "distribution_iso19115_url",  # URL metadato ISO 19115
    "distribution_wfs_url": "distribution_wfs_url",  # URL servicio WFS
}

DEFAULT_DATASET_SHEET      = "dataset"
DEFAULT_DISTRIBUTION_SHEET = "distribution"

DATASET_REQUIRED      = {}
DISTRIBUTION_REQUIRED = {"url"}

# Columnas del mapping V1 que pueden no estar en el Excel (opcionales)
# La validación de headers será permisiva para estas.
ANDINO_V1_DATASET_OPTIONAL  = set()
ANDINO_V1_RESOURCE_OPTIONAL = set()


class XLSXHarvesterClaude(HarvesterBase):
    """
    Harvester para archivos Excel (.xlsx) con el mismo comportamiento que
    CKANHarvester: mismas transformaciones de campos, mapeo de valores y
    defaults via modify_package_dict().

    Estructura esperada del archivo:
      - Pestaña "dataset"      : una fila por dataset; encabezados = display_property DCAT-AP
      - Pestaña "distribution" : una fila por recurso; incluye dataset_identifier (FK)

    Configuración opcional (JSON en el campo "Configuration" del formulario):
    {
        "dataset_sheet":      "dataset",
        "distribution_sheet": "distribution",
        "skip_rows":          0,
        "default_owner_org":  "mi-org",
        "default_tags":       ["tag1", "tag2"],
        "default_extras":     {"clave": "valor"},
        "override_extras":    false,
        "user_agent":         "mi-portal/1.0",
        "force_all":          false
    }

    Nota: default_tags es una lista de strings (nombres de tags), no de dicts.
    """

    config = None

    # ------------------------------------------------------------------ info --

    def info(self):
        return {
            "name": "xlsx_claude",
            "title": "XLSX (Claude)",
            "description": (
                "Harvests remote XLSX files following the datos.gob.ar DCAT-AP schema. "
                "Applies the same field transformations as the CKAN harvester. "
                "Expects a 'dataset' sheet and a 'distribution' sheet."
            ),
            "form_config_interface": "Text",
        }

    # --------------------------------------------------------- validate_config --

    def validate_config(self, config_str):
        if not config_str:
            return config_str
        try:
            cfg = json.loads(config_str)
        except ValueError as e:
            raise ValueError("La configuración debe ser un JSON válido: %s" % e)

        for key in ("dataset_sheet", "distribution_sheet", "default_owner_org", "user_agent"):
            if key in cfg and not isinstance(cfg[key], str):
                raise ValueError("%s debe ser un string" % key)
        if "skip_rows" in cfg and not isinstance(cfg["skip_rows"], int):
            raise ValueError("skip_rows debe ser un entero")
        if "default_tags" in cfg:
            if not isinstance(cfg["default_tags"], list):
                raise ValueError("default_tags debe ser una lista")
            if cfg["default_tags"] and not isinstance(cfg["default_tags"][0], dict):
                raise ValueError("default_tags debe ser una lista de diccionarios")
        if "default_extras" in cfg and not isinstance(cfg["default_extras"], dict):
            raise ValueError("default_extras debe ser un diccionario")
        for key in ("read_only", "force_all", "override_extras"):
            if key in cfg and not isinstance(cfg[key], bool):
                raise ValueError("%s debe ser booleano" % key)

        return config_str

    # ------------------------------------------------------ get_original_url --

    def get_original_url(self, harvest_object_id):
        obj = HarvestObject.get(harvest_object_id)
        return obj.source.url if obj else None

    # ------------------------------------------------------- fields_options --

    @property
    def fields_options(self):
        if not hasattr(self, "_field_options"):
            self._field_options = None
        if self._field_options:
            return self._field_options
        path = os.path.join(os.path.dirname(__file__), "../assets/fields_options.json")
        try:
            with open(path, "r") as f:
                self._field_options = json.load(f)
        except Exception as e:
            log.warning("Error cargando fields_options.json en %s: %s", path, e)
            self._field_options = {}
        return self._field_options

    def get_field_options(self, field_name):
        return self.fields_options.get(field_name, {})

    # ---------------------------------------------------------- gather_stage --

    def _validate_mapping_headers(self, workbook, sheet_name, mapping, skip=0, optional_keys=None):
        """
        Valida que los headers requeridos del mapping existan en la hoja Excel.
        Los headers en optional_keys se ignoran si no están presentes.

        Para V2 (optional_keys=None) la validación es estricta.
        Para V1 se puede pasar un set de keys opcionales.
        """
        optional_keys = {k.lower() for k in (optional_keys or set())}

        try:
            ws = workbook[sheet_name]
        except KeyError:
            raise ValueError("No existe la hoja '%s'" % sheet_name)

        header_row = next(
            ws.iter_rows(
                min_row=skip + 1,
                max_row=skip + 1,
                values_only=True,
            )
        )

        excel_headers = {
            str(h).strip().lower()
            for h in header_row
            if h is not None
        }

        required_headers = {
            str(h).strip().lower()
            for h in mapping.keys()
            if str(h).strip().lower() not in optional_keys
        }

        missing = sorted(required_headers - excel_headers)

        if missing:
            raise ValueError(
                "Faltan columnas requeridas en hoja '%s': %s"
                % (sheet_name, ", ".join(missing))
            )

    def gather_stage(self, harvest_job):
        log.info("XLSXHarvesterClaude gather_stage: %s", harvest_job.source.url)

        self._set_config(harvest_job.source.config)
        source_url = harvest_job.source.url.strip()

        try:
            wb = self._fetch_workbook(source_url)
        except Exception as e:
            self._save_gather_error(
                "No se pudo descargar/leer el XLSX: %s" % e,
                harvest_job,
            )
            return []

        skip     = self.config.get("skip_rows", 0)
        ds_sheet = self.config.get("dataset_sheet", DEFAULT_DATASET_SHEET)
        dist_sheet = self.config.get("distribution_sheet", DEFAULT_DISTRIBUTION_SHEET)

        ds_rows = dist_rows = None
        errors = []

        for ds_map, dist_map, version, ds_optional, dist_optional in [
            (
                ANDINO_V2_DATASET_COLUMN_MAP,
                ANDINO_V2_RESOURCE_COLUMN_MAP,
                "V2",
                set(),   # V2: validación estricta, sin opcionales
                set(),
            ),
            (
                ANDINO_V1_DATASET_COLUMN_MAP,
                ANDINO_V1_RESOURCE_COLUMN_MAP,
                "V1",
                ANDINO_V1_DATASET_OPTIONAL,   # V1: columnas que pueden faltar
                ANDINO_V1_RESOURCE_OPTIONAL,
            ),
        ]:
            try:
                self._validate_mapping_headers(wb, ds_sheet,   ds_map,   skip, ds_optional)
                self._validate_mapping_headers(wb, dist_sheet, dist_map, skip, dist_optional)

                ds_rows_try   = self._read_sheet(wb, ds_sheet,   ds_map,   skip)
                dist_rows_try = self._read_sheet(wb, dist_sheet, dist_map, skip)

            except Exception as e:
                errors.append("Mapeo %s inválido: %s" % (version, e))
                continue

            ds_rows   = ds_rows_try
            dist_rows = dist_rows_try
            log.info("Catálogo leído correctamente con mapeo %s", version)
            break

        if ds_rows is None:
            self._save_gather_error(
                "No se pudo leer el XLSX con ningún mapeo conocido. Detalles: %s"
                % " | ".join(errors),
                harvest_job,
            )
            return []

        # Agrupar distribuciones por dataset_identifier (FK)
        dist_by_dataset = {}
        for d in dist_rows:
            ds_id = str(d.get("id", "")).strip()
            if ds_id:
                dist_by_dataset.setdefault(ds_id, []).append(d)

        # GUIDs de datasets ya harvested en runs anteriores (para detectar eliminados)
        previous_guids = {
            r[0]
            for r in model.Session.query(HarvestObject.guid).filter(
                HarvestObject.harvest_source_id == harvest_job.source.id,
                HarvestObject.current == True,  # noqa: E712
            ).all()
        }

        object_ids = []
        seen_guids = set()

        for i, row in enumerate(ds_rows):
            ds_id = str(row.get("id", "")).strip()

            missing = [
                f for f in DATASET_REQUIRED
                if not str(row.get(f, "")).strip()
            ]
            if missing:
                log.warning("Fila %d omitida: faltan campos %s", i + 2, missing)
                continue

            guid = self._make_guid(ds_id, source_url, i, row)

            if guid in seen_guids:
                log.info("Dataset duplicado descartado: %s", guid)
                continue

            seen_guids.add(guid)
            row["_resources"] = dist_by_dataset.get(ds_id, [])

            content = json.dumps(row, default=str)

            obj = HarvestObject(
                guid=guid,
                job=harvest_job,
                content=content,
                extras=[],
            )
            obj.save()
            object_ids.append(obj.id)

        # Marcar como no-current los datasets que ya no están en el Excel
        deleted_guids = previous_guids - seen_guids
        if deleted_guids:
            log.info(
                "XLSXHarvesterClaude: %d datasets ya no están en el Excel y serán marcados como no-current",
                len(deleted_guids),
            )
            model.Session.query(HarvestObject).filter(
                HarvestObject.guid.in_(deleted_guids),
                HarvestObject.harvest_source_id == harvest_job.source.id,
            ).update({"current": False}, synchronize_session=False)
            model.Session.commit()

        log.info("XLSXHarvesterClaude: %d datasets recolectados", len(object_ids))
        return object_ids

    # ----------------------------------------------------------- fetch_stage --

    def fetch_stage(self, harvest_object):
        # El contenido ya fue serializado en gather_stage; no hay nada que descargar.
        if not harvest_object.content:
            self._save_object_error(
                "El harvest object no tiene contenido", harvest_object, "Fetch"
            )
            return False
        return True

    # ---------------------------------------------------------- import_stage --

    def import_stage(self, harvest_object):
        log.debug("XLSXHarvesterClaude import_stage: %s", harvest_object.id)

        if not harvest_object.content:
            self._save_object_error(
                "Contenido vacío en el objeto de harvest", harvest_object, "Import"
            )
            return False

        self._set_config(harvest_object.source.config)


        # 1. Deserializar la fila guardada en el gather_stage
        try:
            row = json.loads(harvest_object.content)
        except ValueError as e:
            self._save_object_error(
                "JSON inválido en content: %s" % e, harvest_object, "Import"
            )
            return False

        # 2. Construir el package_dict base (campos core de CKAN)
        try:
            package_dict = self._build_package_dict(row, harvest_object)
        except Exception as e:
            self._save_object_error(
                "Error construyendo package_dict: %s" % e, harvest_object, "Import"
            )
            return False

        # 3. Aplicar transformaciones específicas (Theme, SuperTheme, campos temporales, etc.)
        try:
            package_dict = self.modify_package_dict(package_dict, harvest_object)
        except Exception as e:
            self._save_object_error(
                "Error en modify_package_dict: %s" % e, harvest_object, "Import"
            )
            return False

        # 4. Inyectar tags por defecto desde la configuración del harvest source
        # default_tags es una lista de strings (nombres de tags)
        default_tags = self.config.get("default_tags", [])
        if default_tags:
            package_dict.setdefault("tags", [])
            existing_tag_names = {t["name"] for t in package_dict["tags"] if "name" in t}
            for tag_name in default_tags:
                if isinstance(tag_name, str) and tag_name not in existing_tag_names:
                    package_dict["tags"].append({"name": tag_name})

        # 5. Persistir en CKAN
        context = {
            "model": model,
            "session": model.Session,
            "user": self._get_user_name(),
            "ignore_auth": True,
        }

        try:
            exists = model.Package.get(package_dict["id"])

            if exists and exists.state != "deleted":
                log.info("Actualizando dataset existente: %s", package_dict["name"])
                result = get_action("package_update")(context, package_dict)
            else:
                log.info("Creando nuevo dataset: %s", package_dict["name"])
                result = get_action("package_create")(context, package_dict)

            # FIX: actualizar package_id en el HarvestObject para tracking correcto
            harvest_object.package_id = result["id"]
            harvest_object.current = True
            harvest_object.save()

            return True

        except Exception as e:
            self._save_object_error(
                "Error al guardar en CKAN (Acción API): %s" % str(e),
                harvest_object,
                "Import",
            )
            return False

    # ---------------------------------------------------- modify_package_dict --

    def modify_package_dict(self, package_dict, harvest_object):
        """
        Aplica transformaciones y valores por defecto al package_dict para
        alinearlo con el perfil de metadatos V2.

        :param package_dict: dict con campos básicos de CKAN ya incorporados
                             (viene de _build_package_dict)
        :param harvest_object: objeto HarvestObject en curso
        :return: package_dict enriquecido y normalizado
        """

        _BASE_TOP_URL = "http://publications.europa.eu/resource/authority/data-theme/"
        SUPER_THEME_ACRONYM_TO_VALUE = {
            code: _BASE_TOP_URL + code
            for code in [
                "TECH", "ECON", "EDUC", "SOCI", "ENER", "GOVE",
                "JUST", "ENVI", "AGRI", "HEAL", "TRAN", "REGI", "INTR",
            ]
        }
        BASE_GEO_URI = "https://infra.datos.gob.ar/vocabulario/territorio-argentina/"
        ARG_CODES = ['AR','ARG']

        SPATIAL_URIS = [BASE_GEO_URI + s for s in
                        ["Pais/Argentina", "Region/NOA", "Region/NEA", "Region/Cuyo", "Region/Pampeana", "Region/AMBA",
                         "Region/Patagonia", "Provincia/Buenos-Aires", "CABA/Ciudad-Autonoma-de-Buenos-Aires",
                         "Provincia/Catamarca", "Provincia/Chaco", "Provincia/Chubut", "Provincia/Cordoba",
                         "Provincia/Corrientes", "Provincia/Entre-Rios", "Provincia/Formosa", "Provincia/Jujuy",
                         "Provincia/La-Pampa", "Provincia/La-Rioja", "Provincia/Mendoza", "Provincia/Misiones",
                         "Provincia/Neuquen", "Provincia/Rio-Negro", "Provincia/Salta", "Provincia/San-Juan",
                         "Provincia/San-Luis", "Provincia/Santa-Cruz", "Provincia/Santa-Fe",
                         "Provincia/Santiago-del-Estero", "Provincia/Tierra-del-Fuego", "Provincia/Tucuman"]]

        # 1. dataset_description
        if not package_dict.get("dataset_description"):
            package_dict["dataset_description"] = package_dict.get("notes", "")

        # 2. dataset_status — solo se agrega si falta
        if not package_dict.get("dataset_status"):
            package_dict["dataset_status"] = "http://purl.org/adms/status/Completed"

        # 3. dataset_hvdCategory — solo se agrega si falta
        if not package_dict.get("dataset_hvdCategory"):
            package_dict["dataset_hvdCategory"] = "categoria 1"

        #TODO algunos nodos ponen spatial y spatial_uri, ver qué hacer con esos casos
        if not package_dict.get("spatial"):
            package_dict["spatial"] = ""
        spatial_uri = package_dict.get("spatial_uri")
        if spatial_uri:
            if isinstance(spatial_uri, list):
                uri_list = spatial_uri
            else:
                try:
                    parsed = json.loads(spatial_uri)
                    uri_list = parsed if isinstance(parsed, list) else [str(parsed).strip()]
                except (ValueError, TypeError):
                    uri_list = [v.strip() for v in str(spatial_uri).split(",") if v.strip()]

            valid = [u for u in uri_list if u in SPATIAL_URIS]
            package_dict["spatial_uri"] = valid if valid else ""


        # 5. temporal_end — se deriva de temporal_start si contiene "/" (formato ISO 8601 interval)
        if not package_dict.get("temporal_end"):
            temporal_start = str(package_dict.get("temporal_start", "") or "")
            if "/" in temporal_start:
                parts = temporal_start.split("/", 1)
                try:
                    datetime.datetime.strptime(parts[0].strip(), "%Y-%m-%d")
                    datetime.datetime.strptime(parts[1].strip(), "%Y-%m-%d")
                    package_dict["temporal_start"] = parts[0].strip()
                    package_dict["temporal_end"]   = parts[1].strip()
                except ValueError:
                    package_dict["temporal_start"] = ""
                    package_dict["temporal_end"]   = ""

        # 6. dataset_accrualPeriodicity
        if not package_dict.get("dataset_accrualPeriodicity"):
            package_dict["dataset_accrualPeriodicity"] = (
                "http://publications.europa.eu/resource/authority/frequency/CONT"
            )
        else:
            ap_value = package_dict["dataset_accrualPeriodicity"]
            package_dict["dataset_accrualPeriodicity"] = (
                self.get_field_options("accrual_periodicity").get(ap_value, ap_value)
            )

        # 7. dataset_superTheme — sustituir siglas por URLs; default si vacío
        DEFAULT_SUPER_THEME = "Sin tema"
        raw_st = package_dict.get("dataset_superTheme")

        if isinstance(raw_st, str):
            try:
                st_list = json.loads(raw_st)
            except (ValueError, TypeError):
                st_list = [v.strip() for v in raw_st.split(",") if v.strip()]
        elif isinstance(raw_st, list):
            st_list = raw_st
        else:
            st_list = []

        if st_list:
            package_dict["dataset_superTheme"] = [
                SUPER_THEME_ACRONYM_TO_VALUE.get(item, item)
                for item in st_list
            ]
        else:
            package_dict["dataset_superTheme"] = [DEFAULT_SUPER_THEME]

        # 8. dataset_theme — siempre se sobrescribe
        package_dict["dataset_theme"] = "Tema específico 1"

        # 9. dataset_issued / dataset_modified
        if not package_dict.get("dataset_issued"):
            package_dict["dataset_issued"] = package_dict.get("metadata_created", "")
        if not package_dict.get("dataset_modified"):
            package_dict["dataset_modified"] = package_dict.get("metadata_modified", "")

        # 10. Campos faltantes en recursos/distribuciones
        DISTRIBUTION_DEFAULTS = {
            "distribution_character_set": "",
            "distribution_scale":         "",
            "distribution_projection":    "",
            "distribution_iso19115_url":  "",
            "distribution_wfs_url":       ""
        }
        for resource in package_dict.get("resources", []):
            for field, default in DISTRIBUTION_DEFAULTS.items():
                if not resource.get(field):
                    resource[field] = default
            resource['distribution_type'] =  'http://purl.org/dc/dcmitype/Dataset'

        return package_dict

    # ---------------------------------------------------------------- helpers --

    def _set_config(self, config_str):
        try:
            self.config = json.loads(config_str) if config_str else {}
        except ValueError:
            self.config = {}

    def _fetch_workbook(self, url):
        ua = self.config.get("user_agent", "ckanext-harvest-xlsx/1.0")
        resp = requests.get(url, headers={"User-Agent": ua}, timeout=60)
        resp.raise_for_status()
        if "html" in resp.headers.get("Content-Type", ""):
            raise ValueError(
                "La URL devolvió HTML en vez de un archivo XLSX. "
                "Verificá que la URL apunte directamente al archivo."
            )
        return openpyxl.load_workbook(
            io.BytesIO(resp.content), read_only=True, data_only=True
        )

    def _read_sheet(self, wb, sheet_name, column_map, skip=0):
        """Lee una hoja y devuelve lista de dicts mapeados via column_map."""
        if sheet_name not in wb.sheetnames:
            log.warning("Hoja '%s' no encontrada en el workbook", sheet_name)
            return []

        rows = list(wb[sheet_name].iter_rows(values_only=True))
        if not rows:
            return []

        raw_headers = rows[skip]
        headers = [str(h).strip() if h is not None else "" for h in raw_headers]

        result = []
        for row in rows[skip + 1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            raw = {
                headers[j]: (row[j] if j < len(row) else None)
                for j in range(len(headers))
            }
            mapped = {}
            for col_header, val in raw.items():
                field_name = column_map.get(col_header)
                if field_name and val is not None:
                    mapped[field_name] = str(val).strip()
            result.append(mapped)

        return result

    def _make_guid(self, ds_id, source_url, row_index, row):
        """GUID = URL_fuente/dataset_identifier, o MD5 como fallback."""
        if ds_id:
            return "%s/%s" % (source_url.rstrip("/"), ds_id)
        title = row.get("title", row.get("dataset_title", ""))
        raw   = "%s|%s|%d" % (source_url, title, row_index)
        return hashlib.md5(raw.encode("utf-8")).hexdigest()

    def _build_package_dict(self, row, harvest_object):
        """
        Construye el package_dict base con los campos críticos para CKAN:
        title, name, owner_org, tags, resources e id.
        """
        pkg = dict(row)

        if not pkg.get("title"):
            raise ValueError("El dataset no tiene título válido para CKAN")

        # FIX: preservar el name existente en updates para evitar conflictos
        existing_id = self._get_existing_package_id(harvest_object)
        if existing_id:
            try:
                existing = toolkit.get_action("package_show")(
                    {"ignore_auth": True}, {"id": existing_id}
                )
                pkg["name"] = existing["name"]
            except toolkit.ObjectNotFound:
                log.warning(
                    "Package %s ya no existe, se generará un name nuevo", existing_id
                )
                pkg["name"] = self._gen_new_name(pkg.get("name") or pkg["title"])
        else:
            pkg["name"] = self._gen_new_name(pkg.get("name") or pkg["title"])

        # Organización dueña
        if not pkg.get("owner_org"):
            default_org = self.config.get("default_owner_org", "")
            if not default_org:
                source_dataset = toolkit.get_action("package_show")(
                    {"ignore_auth": True}, {"id": harvest_object.source.id}
                )
                default_org = source_dataset.get("owner_org", "")
            if default_org:
                pkg["owner_org"] = default_org

        # Keywords → Tags
        if "dataset_keywords" in pkg and not pkg.get("tags"):
            keywords_raw = pkg.pop("dataset_keywords") or ""
            if keywords_raw:
                pkg["tags"] = [
                    {"name": t.strip()}
                    for t in keywords_raw.split(",")
                    if t.strip()
                ]

        # Trazabilidad del harvest
        pkg["harvest_source_url"] = harvest_object.source.url

        # Procesar distribuciones (_resources → resources)
        resources_raw  = pkg.pop("_resources", [])
        resources_list = []

        for dist in resources_raw:
            res_url = (dist.get("distribution_download_url") or dist.get("url") or "").strip()
            if not res_url:
                continue

            res = {
                "url":         res_url,
                "name":        dist.get("name") or dist.get("distribution_name") or pkg["title"],
                "format":      (dist.get("distribution_format") or dist.get("format") or "").strip().upper(),
                "description": dist.get("distribution_description") or dist.get("description", ""),
            }

            # Campos extendidos de la distribución
            for clave, valor in dist.items():
                if clave not in ("url", "distribution_download_url", "name", "distribution_name",
                                 "format", "distribution_format", "description",
                                 "distribution_description", "id"):
                    val_str = str(valor or "").strip()
                    if val_str:
                        res[clave] = val_str

            resources_list.append(res)

        # Recurso de auxilio si no hay distribuciones válidas
        if not resources_list:
            resources_list = [{
                "url":    harvest_object.source.url,
                "format": "XLSX",
                "name":   "Archivo fuente XLSX",
            }]

        pkg["resources"] = resources_list

        # ID único consistente
        if existing_id:
            pkg["id"] = existing_id
        else:
            pkg["id"] = pkg.get("id") or str(uuid.uuid4())

        return pkg

    def _format_super_theme(self, raw_value):
        """
        Convierte el valor de superTheme del XLSX al formato de lista JSON
        que espera modify_package_dict (ej. '["AGRI"]').
        Acepta:
          - "AGRI"           → '["AGRI"]'
          - "AGRI, SOCI"     → '["AGRI", "SOCI"]'
          - '["AGRI"]'       → '["AGRI"]'  (ya está en formato correcto)
        """
        if not raw_value:
            return ""
        raw_value = raw_value.strip()
        if raw_value.startswith("["):
            return raw_value
        items = [v.strip() for v in raw_value.split(",") if v.strip()]
        return json.dumps(items)

    def _get_existing_package_id(self, harvest_object):
        """
        Devuelve el ID interno del package si el dataset ya fue harvested,
        o None si es la primera vez.
        """
        if harvest_object.package_id:
            return harvest_object.package_id

        previous = (
            HarvestObject.Session.query(HarvestObject)
            .filter(HarvestObject.guid == harvest_object.guid)
            .filter(HarvestObject.package_id != None)   # noqa: E711
            .filter(HarvestObject.id != harvest_object.id)
            .order_by(HarvestObject.gathered.desc())
            .first()
        )
        if previous and previous.package_id:
            try:
                toolkit.get_action("package_show")(
                    {"ignore_auth": True}, {"id": previous.package_id}
                )
                return previous.package_id
            except toolkit.ObjectNotFound:
                log.warning(
                    "Package %s ya no existe en CKAN, se creará uno nuevo",
                    previous.package_id,
                )

        return None

    def _save_gather_error(self, message, harvest_job):
        log.error("XLSXHarvesterClaude gather error: %s", message)
        HarvestGatherError(message=message, job=harvest_job).save()
